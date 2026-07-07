"""Read the Excel sheet into ``Client`` objects, stopping at the first empty row.

Excel layout (see ``config``):
  * Row 1 is a header and is skipped.
  * Column 1 = client name (used in the personalized message).
  * Column 2 = phone number without a country code.
  * Any other columns are ignored for name/phone but exposed for templating.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .. import config
from .models import Client, SkippedRow
from .phone import normalize_phone


def load_clients(
    data_file: Path = config.DATA_FILE,
) -> tuple[list[Client], list[SkippedRow]]:
    """Read the workbook and return ``(valid_clients, skipped_rows)``.

    Only the first worksheet is read. Column 1 is the name and column 2 is the
    phone number; any other columns are ignored. Reading stops at the first
    empty row (see ``config.STOP_AT_FIRST_EMPTY_ROW``). The header row is still
    used to populate each client's ``fields`` map so templates can reference a
    column by its header (e.g. ``{DUE DATE}``).
    """
    if not Path(data_file).exists():
        raise FileNotFoundError(f"Client data file not found: {data_file}")

    workbook = load_workbook(filename=str(data_file), read_only=True, data_only=True)
    sheet = workbook.active

    clients: list[Client] = []
    skipped: list[SkippedRow] = []

    try:
        rows = sheet.iter_rows(values_only=True)

        # Read the header row so we can map column headers to values.
        headers: list[str] = []
        for _ in range(config.HEADER_ROWS):
            header_row = next(rows, None)
            if header_row is not None:
                headers = [str(h).strip() if h is not None else "" for h in header_row]

        # Fixed columns: 1 = name, 2 = phone. Other columns are ignored.
        name_idx = config.NAME_COLUMN - 1
        phone_idx = config.PHONE_COLUMN - 1

        for offset, values in enumerate(rows):
            row_number = config.HEADER_ROWS + 1 + offset

            name_cell = (
                values[name_idx] if values and len(values) > name_idx else None
            )
            phone_cell = (
                values[phone_idx] if values and len(values) > phone_idx else None
            )

            name = str(name_cell).strip() if name_cell is not None else ""
            raw_phone = str(phone_cell).strip() if phone_cell is not None else ""

            # First empty row (blank name AND phone) marks the end of the list.
            if not name and not raw_phone:
                if config.STOP_AT_FIRST_EMPTY_ROW:
                    break
                continue

            if not name:
                skipped.append(SkippedRow(row_number, name, raw_phone, "missing name"))
                continue

            normalized = normalize_phone(raw_phone)
            if normalized is None:
                skipped.append(
                    SkippedRow(row_number, name, raw_phone, "invalid phone number")
                )
                continue

            # Build the header -> value map for templating.
            fields: dict[str, str] = {}
            for col_idx, header in enumerate(headers):
                if not header:
                    continue
                cell = values[col_idx] if len(values) > col_idx else None
                fields[header] = str(cell).strip() if cell is not None else ""
            # Guarantee a NAME field even if the header text differs.
            fields.setdefault("NAME", name)

            clients.append(
                Client(
                    name=name,
                    raw_phone=raw_phone,
                    phone=normalized,
                    row=row_number,
                    fields=fields,
                )
            )
    finally:
        workbook.close()

    return clients, skipped
